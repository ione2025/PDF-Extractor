import os
import json
import base64
import time
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file, Response
from werkzeug.utils import secure_filename
import PyPDF2
import pdfplumber
import fitz  # PyMuPDF
from PIL import Image
import io
import google.generativeai as genai
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pytesseract
from pdf2image import convert_from_path
import threading

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'output'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size
ALLOWED_EXTENSIONS = {'pdf'}

# Gemini API Configuration - Use environment variable for security
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY', 'AIzaSyDLumkxN_6uKWwqJKs5QwOT8jP9sGCW0hQ')
genai.configure(api_key=GEMINI_API_KEY)

# Multi-language OCR configuration
# Supports English, Arabic, Chinese Traditional, Chinese Simplified
TESSERACT_LANGUAGES = 'eng+ara+chi_tra+chi_sim'

# Create necessary directories
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

# Product categories
CATEGORIES = ['Gate', 'Door', 'Fence', 'Handrail', 'Window Protection']

# Global progress tracking
progress_data = {}


def allowed_file(filename):
    """Check if the file has an allowed extension."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def update_progress(task_id, current, total, message, eta_seconds=None):
    """Update progress for a task."""
    progress_data[task_id] = {
        'current': current,
        'total': total,
        'percentage': int((current / total) * 100) if total > 0 else 0,
        'message': message,
        'eta_seconds': eta_seconds,
        'timestamp': time.time()
    }


def perform_ocr_on_pdf(pdf_path, task_id=None):
    """
    Perform multi-language OCR on PDF pages to extract text.
    Supports English, Arabic, Chinese Traditional, and Chinese Simplified.
    Returns concatenated text from all pages.
    """
    ocr_text = ""
    
    try:
        # Convert PDF to images
        if task_id:
            update_progress(task_id, 0, 100, "Converting PDF to images for multi-language OCR...")
        
        images = convert_from_path(pdf_path, dpi=300)
        total_pages = len(images)
        
        for i, image in enumerate(images):
            if task_id:
                update_progress(task_id, i, total_pages, f"OCR scanning page {i+1}/{total_pages} (English/Arabic/Chinese)...")
            
            # Perform OCR with multi-language support
            page_text = pytesseract.image_to_string(image, lang=TESSERACT_LANGUAGES)
            ocr_text += f"\n--- Page {i + 1} (Multi-language OCR) ---\n"
            ocr_text += page_text
        
        return ocr_text
        
    except Exception as e:
        return f"OCR failed: {str(e)}. Please ensure Tesseract language packs are installed (eng, ara, chi_tra, chi_sim)."


def extract_text_pypdf2(pdf_path, task_id=None):
    """Extract text from PDF using PyPDF2."""
    text = ""
    try:
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            num_pages = len(pdf_reader.pages)
            
            for page_num in range(num_pages):
                if task_id:
                    update_progress(task_id, page_num, num_pages, f"Extracting text from page {page_num+1}/{num_pages}...")
                
                page = pdf_reader.pages[page_num]
                text += f"\n--- Page {page_num + 1} ---\n"
                text += page.extract_text()
    except Exception as e:
        raise Exception(f"PyPDF2 extraction failed: {str(e)}")
    
    return text


def extract_text_pdfplumber(pdf_path, task_id=None):
    """Extract text from PDF using pdfplumber."""
    text = ""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            num_pages = len(pdf.pages)
            
            for page_num, page in enumerate(pdf.pages):
                if task_id:
                    update_progress(task_id, page_num, num_pages, f"Extracting text from page {page_num+1}/{num_pages}...")
                
                text += f"\n--- Page {page_num + 1} ---\n"
                page_text = page.extract_text()
                if page_text:
                    text += page_text
    except Exception as e:
        raise Exception(f"pdfplumber extraction failed: {str(e)}")
    
    return text


def extract_images_from_pdf(pdf_path, output_base_folder, task_id=None):
    """
    Extract high-quality images from PDF using PyMuPDF at 300-400 DPI.
    Returns list of extracted image info.
    """
    extracted_images = []
    
    try:
        # Open PDF with PyMuPDF
        pdf_document = fitz.open(pdf_path)
        total_pages = len(pdf_document)
        
        if task_id:
            update_progress(task_id, 0, total_pages, "Extracting images from PDF...")
        
        # High-resolution scaling matrix for 300-400 DPI
        zoom = 4.0  # This gives approximately 300-400 DPI
        mat = fitz.Matrix(zoom, zoom)
        
        for page_num in range(total_pages):
            if task_id:
                update_progress(task_id, page_num, total_pages, f"Extracting images from page {page_num+1}/{total_pages}...")
            
            page = pdf_document[page_num]
            
            # Method 1: Extract embedded images
            image_list = page.get_images(full=True)
            
            for img_index, img in enumerate(image_list):
                xref = img[0]
                base_image = pdf_document.extract_image(xref)
                image_bytes = base_image["image"]
                image_ext = base_image["ext"]
                
                # Convert to PIL Image for processing
                image = Image.open(io.BytesIO(image_bytes))
                
                # Store temporarily for AI analysis
                temp_path = os.path.join(app.config['UPLOAD_FOLDER'], f'temp_p{page_num}_i{img_index}.png')
                image.save(temp_path, 'PNG')
                
                extracted_images.append({
                    'page': page_num + 1,
                    'index': img_index,
                    'path': temp_path,
                    'type': 'embedded'
                })
            
            # Method 2: Render page as high-res image (fallback if no embedded images)
            if not image_list:
                pix = page.get_pixmap(matrix=mat)
                img_data = pix.tobytes("png")
                
                temp_path = os.path.join(app.config['UPLOAD_FOLDER'], f'temp_p{page_num}_rendered.png')
                with open(temp_path, 'wb') as f:
                    f.write(img_data)
                
                extracted_images.append({
                    'page': page_num + 1,
                    'index': 0,
                    'path': temp_path,
                    'type': 'rendered'
                })
        
        pdf_document.close()
        return extracted_images
        
    except Exception as e:
        raise Exception(f"Image extraction failed: {str(e)}")


def analyze_image_with_gemini(image_path):
    """
    Analyze image using Gemini 1.5 Flash API to extract:
    - SKU number
    - Product category
    - Product description
    - SVG path for silhouette
    - Primary and secondary colors
    """
    try:
        # Initialize Gemini model
        model = genai.GenerativeModel('gemini-1.5-flash')
        
        # Read and encode image
        with open(image_path, 'rb') as f:
            image_data = f.read()
        
        # Prepare image for API
        image_parts = [{
            'mime_type': 'image/png',
            'data': image_data
        }]
        
        # Craft the prompt
        prompt = """Analyze this product image and provide the following information in JSON format:

1. "sku": The SKU number printed inside this image. If no SKU is found, return "Unknown".
2. "category": Determine the product category from [Gate, Door, Fence, Handrail, Window Protection]. If unclear, return "Unknown".
3. "description": Provide a brief product description (1-2 sentences) describing what the product is, its features, or purpose. If no clear description can be determined, return "No description available".
4. "svg_path": Generate a clean SVG path string that represents the product's silhouette for 3D extrusion. This should be a simplified outline of the main product shape.
5. "primary_color": The primary hex color of the product (e.g., "#000000").
6. "secondary_color": The secondary hex color of the product (e.g., "#FFFFFF").

Return ONLY a valid JSON object with these exact keys. Example:
{
    "sku": "ABC123",
    "category": "Gate",
    "description": "Modern sliding gate with decorative panels and reinforced frame",
    "svg_path": "M 10,10 L 100,10 L 100,100 L 10,100 Z",
    "primary_color": "#2C3E50",
    "secondary_color": "#ECF0F1"
}"""
        
        # Send request to Gemini
        response = model.generate_content([prompt, image_parts[0]])
        
        # Parse response
        response_text = response.text.strip()
        
        # Try to extract JSON from response
        # Sometimes the API wraps JSON in markdown code blocks
        if '```json' in response_text:
            response_text = response_text.split('```json')[1].split('```')[0].strip()
        elif '```' in response_text:
            response_text = response_text.split('```')[1].split('```')[0].strip()
        
        analysis = json.loads(response_text)
        
        return analysis
        
    except Exception as e:
        # Return default values on error
        return {
            'sku': 'Unknown',
            'category': 'Unknown',
            'description': 'No description available',
            'svg_path': '',
            'primary_color': '#000000',
            'secondary_color': '#FFFFFF',
            'error': str(e)
        }


def save_image_and_metadata(image_path, analysis, output_base_folder):
    """
    Save extracted image as WebP and create JSON metadata file.
    Organize into category folders.
    """
    sku = analysis.get('sku', 'Unknown')
    category = analysis.get('category', 'Unknown')
    
    # Skip if no SKU found
    if sku == 'Unknown' or not sku:
        return None
    
    # Create category folder
    category_folder = os.path.join(output_base_folder, category)
    os.makedirs(category_folder, exist_ok=True)
    
    # Generate safe filename
    safe_sku = secure_filename(sku)
    
    # Save image as WebP
    image = Image.open(image_path)
    webp_path = os.path.join(category_folder, f'{safe_sku}.webp')
    image.save(webp_path, 'WEBP', quality=95)
    
    # Save JSON metadata
    json_path = os.path.join(category_folder, f'{safe_sku}.json')
    metadata = {
        'sku': sku,
        'category': category,
        'description': analysis.get('description', 'No description available'),
        'svg_path': analysis.get('svg_path', ''),
        'primary_color': analysis.get('primary_color', '#000000'),
        'secondary_color': analysis.get('secondary_color', '#FFFFFF')
    }
    
    with open(json_path, 'w') as f:
        json.dump(metadata, f, indent=2)
    
    return {
        'sku': sku,
        'category': category,
        'description': analysis.get('description', 'No description available'),
        'webp_path': webp_path,
        'json_path': json_path
    }


def create_excel_report(products_data, output_folder, pdf_name):
    """
    Create an Excel file with SKU codes and product descriptions.
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Product SKUs"
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Header style
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    headers = ['SKU', 'Category', 'Description', 'Primary Color', 'Secondary Color']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data
    for row_idx, product in enumerate(products_data, start=2):
        ws.cell(row=row_idx, column=1, value=product.get('sku', ''))
        ws.cell(row=row_idx, column=2, value=product.get('category', ''))
        ws.cell(row=row_idx, column=3, value=product.get('description', ''))
        ws.cell(row=row_idx, column=4, value=product.get('primary_color', ''))
        ws.cell(row=row_idx, column=5, value=product.get('secondary_color', ''))
        
        # Align cells
        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical="top")
    
    # Save Excel file
    excel_path = os.path.join(output_folder, f'{pdf_name}_SKU_Report.xlsx')
    wb.save(excel_path)
    
    return excel_path


@app.route('/')
def index():
    """Render the main page."""
    return render_template('index.html')


@app.route('/extract', methods=['POST'])
def extract():
    """Handle PDF upload and text extraction with OCR."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type. Only PDF files are allowed.'}), 400
    
    try:
        # Save the uploaded file
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Generate task ID for progress tracking
        task_id = f"text_{int(time.time())}"
        
        # Get extraction method from request (default to pdfplumber)
        method = request.form.get('method', 'pdfplumber')
        use_ocr = request.form.get('use_ocr', 'true').lower() == 'true'
        
        # Extract text based on selected method
        if method == 'pypdf2':
            extracted_text = extract_text_pypdf2(filepath, task_id)
        else:
            extracted_text = extract_text_pdfplumber(filepath, task_id)
        
        # Perform OCR if enabled and text is minimal
        if use_ocr and len(extracted_text.strip()) < 100:
            update_progress(task_id, 50, 100, "Performing OCR scan...")
            ocr_text = perform_ocr_on_pdf(filepath, task_id)
            extracted_text += "\n\n=== OCR RESULTS ===\n" + ocr_text
        
        # Clean up - delete the uploaded file after extraction
        os.remove(filepath)
        
        return jsonify({
            'success': True,
            'text': extracted_text,
            'filename': filename,
            'method': method,
            'ocr_used': use_ocr
        })
    
    except Exception as e:
        # Clean up file if it exists
        if os.path.exists(filepath):
            os.remove(filepath)
        
        return jsonify({
            'error': f'Extraction failed: {str(e)}'
        }), 500


@app.route('/extract-images', methods=['POST'])
def extract_images():
    """Handle PDF upload and image extraction with AI analysis."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type. Only PDF files are allowed.'}), 400
    
    try:
        # Save the uploaded file
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Create task ID for progress tracking
        task_id = f"images_{int(time.time())}"
        
        # Create output folder for this PDF
        pdf_name = os.path.splitext(filename)[0]
        output_folder = os.path.join(app.config['OUTPUT_FOLDER'], pdf_name)
        os.makedirs(output_folder, exist_ok=True)
        
        # Step 1: Extract images from PDF
        start_time = time.time()
        extracted_images = extract_images_from_pdf(filepath, output_folder, task_id)
        
        if not extracted_images:
            os.remove(filepath)
            return jsonify({
                'error': 'No images found in PDF'
            }), 400
        
        # Step 2: Analyze each image with Gemini AI
        results = []
        products_data = []
        skipped = 0
        processed = 0
        total_images = len(extracted_images)
        
        for idx, img_info in enumerate(extracted_images):
            try:
                # Calculate ETA
                if idx > 0:
                    elapsed = time.time() - start_time
                    avg_time_per_image = elapsed / idx
                    remaining_images = total_images - idx
                    eta_seconds = int(avg_time_per_image * remaining_images)
                else:
                    eta_seconds = None
                
                # Update progress
                update_progress(
                    task_id, 
                    idx, 
                    total_images, 
                    f"Analyzing image {idx+1}/{total_images} with AI...",
                    eta_seconds
                )
                
                # Analyze image
                analysis = analyze_image_with_gemini(img_info['path'])
                
                # Save if SKU is found
                if analysis.get('sku') != 'Unknown' and analysis.get('sku'):
                    save_result = save_image_and_metadata(img_info['path'], analysis, output_folder)
                    
                    if save_result:
                        results.append({
                            'page': img_info['page'],
                            'sku': analysis['sku'],
                            'category': analysis['category'],
                            'description': analysis.get('description', ''),
                            'saved': True
                        })
                        
                        products_data.append({
                            'sku': analysis['sku'],
                            'category': analysis['category'],
                            'description': analysis.get('description', 'No description available'),
                            'primary_color': analysis.get('primary_color', '#000000'),
                            'secondary_color': analysis.get('secondary_color', '#FFFFFF')
                        })
                        
                        processed += 1
                    else:
                        skipped += 1
                else:
                    skipped += 1
                
                # Clean up temp image
                if os.path.exists(img_info['path']):
                    os.remove(img_info['path'])
                    
            except Exception as e:
                print(f"Error processing image: {str(e)}")
                skipped += 1
                continue
        
        # Step 3: Create Excel report
        excel_path = None
        if products_data:
            update_progress(task_id, total_images, total_images, "Creating Excel report...")
            excel_path = create_excel_report(products_data, output_folder, pdf_name)
        
        # Clean up uploaded PDF
        os.remove(filepath)
        
        # Clear progress data
        if task_id in progress_data:
            del progress_data[task_id]
        
        return jsonify({
            'success': True,
            'filename': filename,
            'total_images': len(extracted_images),
            'processed': processed,
            'skipped': skipped,
            'results': results,
            'output_folder': output_folder,
            'excel_file': os.path.basename(excel_path) if excel_path else None
        })
    
    except Exception as e:
        # Clean up files if they exist
        if os.path.exists(filepath):
            os.remove(filepath)
        
        return jsonify({
            'error': f'Image extraction failed: {str(e)}'
        }), 500


@app.route('/progress/<task_id>')
def get_progress(task_id):
    """Get progress for a specific task."""
    if task_id in progress_data:
        return jsonify(progress_data[task_id])
    else:
        return jsonify({'error': 'Task not found'}), 404


@app.route('/download-excel/<filename>')
def download_excel(filename):
    """Download Excel report."""
    try:
        # Security: Use secure_filename to prevent path traversal
        safe_filename = secure_filename(filename)
        
        # Construct full path
        file_path = os.path.join(app.config['OUTPUT_FOLDER'], safe_filename)
        
        # Verify the file is within OUTPUT_FOLDER (prevent directory traversal)
        abs_output_folder = os.path.abspath(app.config['OUTPUT_FOLDER'])
        abs_file_path = os.path.abspath(file_path)
        
        if not abs_file_path.startswith(abs_output_folder):
            return jsonify({'error': 'Invalid file path'}), 403
        
        if os.path.exists(abs_file_path):
            return send_file(abs_file_path, as_attachment=True)
        else:
            return jsonify({'error': 'File not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/health')
def health():
    """Health check endpoint."""
    return jsonify({'status': 'healthy'})


if __name__ == '__main__':
    # Get configuration from environment variables
    import os
    debug_mode = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    port = int(os.environ.get('PORT', 5000))
    host = os.environ.get('HOST', '0.0.0.0')
    
    print("=" * 60)
    print("PDF Extractor Pro - Web Application")
    print("=" * 60)
    print(f"\n🌐 Server starting on http://{host}:{port}")
    print(f"📊 Debug mode: {debug_mode}")
    print("\n💡 Access the application by opening the URL in your browser")
    print("🛑 Press CTRL+C to stop the server")
    print("=" * 60)
    print()
    
    app.run(debug=debug_mode, host=host, port=port)
